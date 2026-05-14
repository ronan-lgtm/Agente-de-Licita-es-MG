"""
relatorio.py
============
Lê o SQLite (já classificado), aplica filtros de município e relevância,
e gera o Excel do dia com 5 abas.

Modos:
  python relatorio.py           → novidades das últimas 24h
  python relatorio.py --tudo    → tudo no banco (municípios alvo)
  python relatorio.py --dias 7  → últimos 7 dias
"""

import sqlite3
import pandas as pd
import sys
import logging
from pathlib import Path
from datetime import datetime, timedelta, date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH     = Path(r"C:\Users\User\Desktop\Agente de Licitações\licitacoes_mg.db")
OUTPUT_DIR  = Path(r"C:\Users\User\Desktop\Agente de Licitações")
HOJE        = date.today()

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("relatorio")

# ── Municípios alvo (IBGE) ────────────────────────────────────────────────────
# Adicione ou remova conforme sua carteira

IBGE_ALVO: dict[str, str] = {
    "3106200": "Belo Horizonte",   "3106705": "Betim",
    "3117876": "Contagem",         "3121605": "Divinópolis",
    "3127701": "Governador Valadares", "3131307": "Ipatinga",
    "3132404": "Itabira",          "3144805": "Itabirito",
    "3136702": "Juiz de Fora",     "3140001": "Mariana",
    "3143302": "Montes Claros",    "3143906": "Nova Lima",
    "3144656": "Nova Serrana",     "3146107": "Ouro Preto",
    "3147006": "Paracatu",         "3152501": "Poços de Caldas",
    "3152006": "Ponte Nova",       "3156700": "Sabará",
    "3157807": "Santa Bárbara",    "3162500": "São João del Rei",
    "3164431": "São Sebastião do Paraíso", "3168705": "Timóteo",
    "3170107": "Uberaba",          "3118601": "Congonhas",
    "3118304": "Conselheiro Lafaiete",    "3117504": "Conceição do Mato Dentro",
    "3119401": "Coronel Fabriciano",      "3109006": "Brumadinho",
    "3122306": "Extrema",          "3128600": "Grão Mogol",
    "3135050": "Janaúba",          "3135209": "Jaíba",
    "3137601": "Leopoldina",       "3139409": "Manhuaçu",
    "3140654": "Matozinhos",       "3107109": "Bom Despacho",
    "3107505": "Boa Esperança",    "3108503": "Botumirim",
    "3109204": "Buenópolis",       "3109402": "Buritizeiro",
    "3111200": "Campo Belo",       "3112703": "Capitão Enéas",
    "3118809": "Coração de Jesus", "3119104": "Corinto",
    "3121605": "Curvelo",          "3122355": "Diamantina",
    "3126901": "Francisco Sá",     "3128105": "Glaucilândia",
    "3134202": "Ituiutaba",        "3132503": "Itamarandiba",
    "3137536": "Lagoa Santa",      "3141801": "Mirabela",
    "3142502": "Montalvânia",      "3142700": "Monte Azul",
    "3144953": "Nova Porteirinha", "3145059": "Novorizonte",
    "3149309": "Pedro Leopoldo",   "3151206": "Pirapora",
    "3152105": "Porteirinha",      "3154804": "Riacho dos Machados",
    "3155504": "Rio Pardo de Minas","3156908": "Sabinópolis",
    "3157203": "Salinas",          "3166105": "Serro",
    "3168309": "Taiobeiras",       "3169703": "Turmalina",
    "3170529": "Urucuia",          "3171709": "Várzea da Palma",
    "3171808": "Varzelândia",      "3103405": "Araçuaí",
    "3110608": "Caeté",            "3115300": "Catas Altas",
    "3117504": "Conceição do Mato Dentro", "3119708": "Coronel Fabriciano",
    "3124807": "Esmeraldas",       "3128709": "Guanhães",
    "3133808": "Itapecerica",      "3134103": "Itatiaiuçu",
    "3134509": "Itaúna",           "3147501": "Pains",
    "3152204": "Prados",           "3155306": "Rio Piracicaba",
    "3157401": "Santa Maria de Itabira",
    # Municípios adicionados — estavam faltando no dicionário original
    "3147105": "Pará de Minas",         "3104502": "Araxá",
    "3105905": "Barra Longa",           "3107000": "Barroso",
    "3106507": "Barão de Cocais",       "3107158": "Belo Vale",
    "3108602": "Brasília de Minas",     "3111002": "Camanducaia",
    "3111150": "Campo Azul",            "3116308": "Chapada Gaúcha",
    "3116902": "Conceição da Barra de Minas", "3119500": "Coromandel",
    "3122900": "Dom Joaquim",           "3124500": "Espinosa",
    "3125606": "Felixlândia",           "3125903": "Ferros",
    "3126109": "Formiga",              "3131000": "Icaraí de Minas",
    "3130903": "Igarapé",              "3132800": "Itacambira",
    "3132909": "Itacarambi",           "3135357": "Jeceaba",
    "3136009": "Joaquim Felício",      "3137502": "Lassance",
    "3143450": "Montezuma",            "3141900": "Morro do Pilar",
    "3144102": "Nazareno",             "3144904": "Novo Cruzeiro",
    "3145901": "Oliveira",             "3147600": "Passabém",
    "3149457": "Pedras de Maria da Cruz", "3152303": "Pratápolis",
    "3153400": "Presidente Kubitschek","3155405": "Rio Espera",
    "3156304": "Rosário da Limeira",   "3158706": "Santo Antônio do Rio Abaixo",
    "3162450": "São Gonçalo do Rio Abaixo", "3163706": "São Joaquim de Bicas",
    "3165560": "São Sebastião da Vargem Alegre", "3165701": "São Sebastião do Rio Preto",
    "3165776": "São Tiago",            "3168804": "Taquaraçu de Minas",
    "3170008": "Ubaí",                 "3171600": "Vargem Grande do Rio Pardo",
    "3171907": "Verdelândia",
}


# ── Estilo Excel ──────────────────────────────────────────────────────────────

COR = {
    "header" : "1F4E79",
    "alta"   : "D9EAD3",
    "media"  : "FFF2CC",
    "baixa"  : "FCE4D6",
    "irrel"  : "F2F2F2",
    "white"  : "FFFFFF",
}

LARGURAS = {
    "Tier":14, "Score":7, "Categoria":22, "Termos Encontrados":30,
    "Modalidade":13, "Município":24, "Alvo":6, "Status":14,
    "Dias Restantes":13, "Objeto":70, "Valor (R$)":18,
    "Data Publicação":15, "Data Abertura":14, "Data Encerramento":18,
    "Órgão":38, "Link PNCP":58,
    # resumo
    "Total":8, "Alta":8, "Média":8, "Em Aberto":10, "Valor Total":20,
}


def formatar_wb(wb):
    thin  = Side(style="thin", color="BFBFBF")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    for aba in wb.sheetnames:
        ws   = wb[aba]
        cols = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

        # Cabeçalho
        for ci in range(1, ws.max_column + 1):
            c = ws.cell(1, ci)
            c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            c.fill      = PatternFill("solid", start_color=COR["header"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = borda
        ws.row_dimensions[1].height = 30

        tier_ci = (cols.index("Tier")          + 1) if "Tier"          in cols else None
        dias_ci = (cols.index("Dias Restantes") + 1) if "Dias Restantes" in cols else None
        val_ci  = (cols.index("Valor (R$)")     + 1) if "Valor (R$)"     in cols else None
        vt_ci   = (cols.index("Valor Total")    + 1) if "Valor Total"    in cols else None
        link_ci = (cols.index("Link PNCP")      + 1) if "Link PNCP"      in cols else None

        for ri in range(2, ws.max_row + 1):
            tv  = str(ws.cell(ri, tier_ci).value or "") if tier_ci else ""
            dv  = ws.cell(ri, dias_ci).value            if dias_ci else None
            cor = (COR["alta"]  if "Alta"  in tv else
                   COR["media"] if "Média" in tv else
                   COR["baixa"] if "Baixa" in tv else COR["irrel"])

            for ci in range(1, ws.max_column + 1):
                c = ws.cell(ri, ci)
                c.font      = Font(name="Arial", size=9)
                c.fill      = PatternFill("solid", start_color=cor if tier_ci else COR["white"])
                c.border    = borda
                c.alignment = Alignment(vertical="center")

            if dias_ci and isinstance(dv, (int, float)) and 0 <= dv <= 3:
                ws.cell(ri, dias_ci).font = Font(name="Arial", size=9, bold=True, color="FF0000")

            for ci_f, fmt in [(val_ci, "R$ #,##0.00"), (vt_ci, "R$ #,##0.00")]:
                if ci_f:
                    c = ws.cell(ri, ci_f)
                    c.number_format = fmt
                    c.alignment = Alignment(horizontal="right", vertical="center")

            if link_ci:
                c = ws.cell(ri, link_ci)
                link_val = str(c.value or "")
                if link_val.startswith("http"):
                    c.hyperlink = link_val
                c.font = Font(name="Arial", size=9, color="0563C1", underline="single")

        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = LARGURAS.get(col or "", 15)

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions


# ── Gera relatório ────────────────────────────────────────────────────────────

def gerar(dias: int = 90, apenas_alvo: bool = False):

    conn   = sqlite3.connect(DB_PATH)

    limite_enc = (datetime.now() - timedelta(days=3)).strftime("%Y-%m-%d")

    query = f"""
        SELECT
            tier            AS "Tier",
            score           AS "Score",
            categoria       AS "Categoria",
            termos_encontrados AS "Termos Encontrados",
            modalidade_nome AS "Modalidade",
            municipio_nome  AS "Município",
            municipio_ibge  AS "IBGE",
            objeto          AS "Objeto",
            valor_estimado  AS "Valor (R$)",
            data_publicacao AS "Data Publicação",
            data_abertura   AS "Data Abertura",
            data_encerramento AS "Data Encerramento",
            orgao_nome      AS "Órgão",
            situacao_nome   AS "Situação",
            link_pncp       AS "Link PNCP",
            data_coleta     AS "Data Coleta"
        FROM licitacoes
        WHERE (
            -- Abertas com pelo menos 1 termo encontrado (score >= 2)
            (data_encerramento >= date('now')
            OR data_encerramento IS NULL
            OR data_encerramento = '')
            AND (score IS NULL OR score >= 2)
        )
        OR (
            -- Encerradas só dos últimos 3 dias, relevantes (score > 0)
            data_encerramento >= '{limite_enc}'
            AND data_encerramento < date('now')
            AND score > 0
        )
        ORDER BY
            CASE WHEN data_encerramento >= date('now') OR data_encerramento IS NULL THEN 0 ELSE 1 END,
            score DESC,
            data_encerramento ASC
    """

    df = pd.read_sql_query(query, conn)
    conn.close()

    if df.empty:
        log.warning("Nenhum resultado encontrado para o período.")
        return

    # Marca municípios alvo
    df["Alvo"] = df["IBGE"].apply(lambda x: "✅" if str(x) in IBGE_ALVO else "🔵")
    df["Município"] = df.apply(
        lambda r: IBGE_ALVO.get(str(r["IBGE"]), r["Município"]), axis=1
    )

    # Status e dias restantes
    def calc_status(enc_str):
        try:
            d = datetime.fromisoformat(enc_str[:19]).date()
            dias_r = (d - HOJE).days
            status = "🟢 Aberta" if dias_r >= 0 else "🔴 Encerrada"
            return status, dias_r, d.strftime("%d/%m/%Y")
        except Exception:
            return "Sem prazo", None, enc_str[:10] if enc_str else ""

    status_data = df["Data Encerramento"].apply(calc_status)
    df["Status"]           = status_data.apply(lambda x: x[0])
    df["Dias Restantes"]   = status_data.apply(lambda x: x[1])
    df["Data Encerramento"]= status_data.apply(lambda x: x[2])

    # Formata data publicação e abertura
    for col in ["Data Publicação", "Data Abertura"]:
        df[col] = df[col].apply(
            lambda x: datetime.fromisoformat(x[:10]).strftime("%d/%m/%Y")
            if x and len(x) >= 10 else ""
        )

    # Colunas finais para o Excel
    COLS = [
        "Tier", "Score", "Categoria", "Termos Encontrados",
        "Modalidade", "Município", "Alvo", "Status", "Dias Restantes",
        "Objeto", "Valor (R$)", "Data Publicação",
        "Data Abertura", "Data Encerramento", "Órgão", "Link PNCP",
    ]

    df = df[COLS]
    df["Valor (R$)"] = pd.to_numeric(df["Valor (R$)"], errors="coerce")

    if apenas_alvo:
        df = df[df["Alvo"] == "✅"].copy()

    df_alta   = df[df["Tier"] == "⭐⭐⭐ Alta"].copy()
    df_am     = df[df["Tier"].isin(["⭐⭐⭐ Alta", "⭐⭐ Média"])].copy()
    df_aberto = df[(df["Status"] == "🟢 Aberta") & (df["Alvo"] == "✅")].copy()
    df_alvo   = df[df["Alvo"] == "✅"].copy()

    resumo = (
        df_alvo.groupby("Município")
        .agg(
            Total      =("Link PNCP", "count"),
            Alta       =("Tier",      lambda x: (x == "⭐⭐⭐ Alta").sum()),
            Média      =("Tier",      lambda x: (x == "⭐⭐ Média").sum()),
            Em_Aberto  =("Status",    lambda x: (x == "🟢 Aberta").sum()),
            Valor_Total=("Valor (R$)","sum"),
        )
        .rename(columns={"Em_Aberto": "Em Aberto", "Valor_Total": "Valor Total"})
        .sort_values(["Alta", "Total"], ascending=False)
        .reset_index()
    )

    # Nome do arquivo com data
    nome = f"licitacoes_mg_{HOJE:%Y%m%d}.xlsx"
    path = OUTPUT_DIR / nome

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_alta.to_excel(  writer, sheet_name="⭐⭐⭐ Alta",          index=False)
        df_am.to_excel(    writer, sheet_name="⭐⭐ Média e Alta",    index=False)
        df_aberto.to_excel(writer, sheet_name="🟢 Em Aberto",        index=False)
        df_alvo.to_excel(  writer, sheet_name="✅ Municípios Alvo",  index=False)
        df.to_excel(       writer, sheet_name="Todos MG",            index=False)
        resumo.to_excel(   writer, sheet_name="📍 Por Município",    index=False)

    wb = load_workbook(path)
    formatar_wb(wb)
    wb.save(path)

    log.info(f"✅ Salvo: {path}")
    log.info(f"   Total MG          : {len(df)}")
    log.info(f"   Municípios alvo   : {len(df_alvo)}")
    log.info(f"   ⭐⭐⭐ Alta          : {len(df_alta)}")
    log.info(f"   🟢 Em aberto (alvo): {len(df_aberto)}")

    if not df_alta.empty:
        log.info("\n🎯 Destaques:")
        for _, r in df_alta.sort_values("Dias Restantes", na_position="last").head(8).iterrows():
            d   = f"{int(r['Dias Restantes']):>3}d" if r["Dias Restantes"] is not None else "  —"
            val = f"R${r['Valor (R$)']:>12,.0f}" if pd.notna(r["Valor (R$)"]) else "          —"
            log.info(f"  ⏰{d} | {val} | {r['Município']:<22} | {str(r['Objeto'])[:50]}...")


if __name__ == "__main__":
    dias      = int(sys.argv[sys.argv.index("--dias") + 1]) if "--dias" in sys.argv else 90
    apenas_alvo = "--alvo" in sys.argv
    gerar(dias=dias, apenas_alvo=apenas_alvo)